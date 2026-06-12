#!/usr/bin/env python3
"""
OMB Public Budget Database outlays extractor (Phase 44, Plan 04).

Usage: python extractOMBPublicBudgetDB.py <outlays.xlsx> <hist03z2.xlsx> <fiscal_year>
Emits JSON to stdout:
  { "function_titles": {"050": "National Defense", ...},
    "rows": [{agency, bureau, account, subfunction_code, subfunction_title,
              function_code, function_title, bea_category, amount}, ...] }
Amounts normalized to DOLLARS (file is in THOUSANDS — verified: FY2025 column
sums to exactly 7,011,105,000 thousands = OMB Hist 1.1 outlays to the thousand).

Function titles come from hist03z2 ("NNN Title:" rows; function codes end in 0,
subfunction codes end in 1-9) — sourced mapping, never from model memory.
"""
import json
import re
import sys
import warnings

warnings.filterwarnings("ignore")
import openpyxl  # noqa: E402


def fatal(msg):
    sys.stderr.write(f"FATAL: {msg}\n")
    sys.exit(1)


def function_titles_from_hist32(path):
    ws = openpyxl.load_workbook(path).active
    titles = {}
    for row in ws.iter_rows(values_only=True):
        m = re.match(r"^(\d{3})\s+(.+?):?\s*$", str(row[0] or "").strip())
        if m and m.group(1).endswith("0"):  # functions end in 0; subfunctions in 1-9
            titles[m.group(1)] = m.group(2).rstrip(":").strip()
    if len(titles) < 18:
        fatal(f"only {len(titles)} function titles found in hist03z2 — layout changed?")
    return titles


def main():
    if len(sys.argv) != 4:
        fatal("Usage: extractOMBPublicBudgetDB.py <outlays.xlsx> <hist03z2.xlsx> <fiscal_year>")
    outlays_path, hist32_path, fy = sys.argv[1], sys.argv[2], sys.argv[3]

    titles = function_titles_from_hist32(hist32_path)

    wb = openpyxl.load_workbook(outlays_path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h or "").strip() for h in next(rows_iter)]

    def col(name):
        if name not in header:
            fatal(f"column '{name}' not in outlays file header")
        return header.index(name)

    c_agency = col("Agency Name")
    c_bureau = col("Bureau Name")
    c_account = col("Account Name")
    c_sfc = col("Subfunction Code")
    c_sft = col("Subfunction Title")
    c_bea = col("BEA Category")
    c_amt = col(fy)

    out = []
    total = 0.0
    for r in rows_iter:
        v = r[c_amt]
        if v in (None, ""):
            continue
        try:
            amount = float(v) * 1000.0  # thousands -> dollars
        except (TypeError, ValueError):
            fatal(f"non-numeric {fy} amount: {v!r} (account {r[c_account]!r})")
        total += amount

        sfc = str(r[c_sfc] or "").strip().zfill(3)
        fc = sfc[:2] + "0"  # function code: subfunction with last digit zeroed
        if fc not in titles:
            fatal(f"function code {fc} (from subfunction {sfc}) missing in hist03z2 titles")
        out.append({
            "agency": str(r[c_agency] or "").strip(),
            "bureau": str(r[c_bureau] or "").strip(),
            "account": str(r[c_account] or "").strip(),
            "subfunction_code": sfc,
            "subfunction_title": str(r[c_sft] or "").strip(),
            "function_code": fc,
            "function_title": titles[fc],
            "bea_category": str(r[c_bea] or "").strip(),
            "amount": amount,
        })

    json.dump({"function_titles": titles, "rows": out, "net_total": total}, sys.stdout)
    sys.stderr.write(f"OK: {len(out)} account rows, net total {total:,.0f} dollars\n")


if __name__ == "__main__":
    main()
